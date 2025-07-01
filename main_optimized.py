import sys
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
from pathlib import Path
import time
import openpyxl
import csv
import re
import traceback
import logging
import glob
import subprocess
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Garante que não tenta abrir janela de plot

logging.basicConfig(
    filename="erro_consolidador.log",
    level=logging.DEBUG,
    format="%(asctime)s %(levelname)s %(message)s"
)

def debug_print(message):
    """Print com timestamp para debug"""
    timestamp = time.strftime('%H:%M:%S')
    print(f"[{timestamp}] {message}")

class ReportProcessor:
    """Classe para processar os relatórios das empresas"""
    
    def __init__(self):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        os.chdir(self.script_dir)
        
    def treat_data_iguatemi(self, filepath, output_path):
        """Processa dados da Iguatemi"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")
        aba_a_manter = 'Indicadores | Indicators'
        temp_csv = os.path.join(self.script_dir, "temp_iguatemi.csv")
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if aba_a_manter not in wb.sheetnames:
            available_sheets = ", ".join(wb.sheetnames)
            raise ValueError(f"Aba '{aba_a_manter}' não encontrada. Abas disponíveis: {available_sheets}")
        ws = wb[aba_a_manter]
        with open(temp_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                new_row = []
                for cell in row:
                    if isinstance(cell, float):
                        cell = str(cell).replace('.', ',')
                    elif isinstance(cell, str):
                        cell = cell.replace('.', ',')
                    new_row.append(cell)
                writer.writerow(new_row)
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = aba_a_manter
        with open(temp_csv, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws_out.append(row)
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb_out.save(output_path)
        if os.path.exists(temp_csv):
            os.remove(temp_csv)
        debug_print(f"Iguatemi processado: {output_path}")

    def treat_data_allos(self, filepath, output_path):
        """Processa dados da Allos"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")
        
        aba_a_manter = 'Indicadores'
        wb = openpyxl.load_workbook(filepath)

        if aba_a_manter not in wb.sheetnames:
            available_sheets = ", ".join(wb.sheetnames)
            raise ValueError(f"Aba '{aba_a_manter}' não encontrada. Abas disponíveis: {available_sheets}")

        # Remover todas as abas que não sejam a desejada
        for aba in wb.sheetnames[:]:
            if aba != aba_a_manter:
                wb.remove(wb[aba])

        # Criar diretório se não existir
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Salvar as alterações
        wb.save(output_path)
        debug_print(f"Allos processado: {output_path}")

    def treat_data_multiplan(self, filepath, output_path):
        """Processa dados da Multiplan"""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")
        
        aba_a_manter = 'Indicadores | Indicators'
        temp_csv = os.path.join(self.script_dir, "temp_multiplan.csv")

        # Extrair aba para CSV
        wb = openpyxl.load_workbook(filepath, data_only=True)
        
        if aba_a_manter not in wb.sheetnames:
            available_sheets = ", ".join(wb.sheetnames)
            raise ValueError(f"Aba '{aba_a_manter}' não encontrada. Abas disponíveis: {available_sheets}")
        
        ws = wb[aba_a_manter]
        with open(temp_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                # Troca pontos por vírgulas em strings e floats
                new_row = []
                for cell in row:
                    if isinstance(cell, float):
                        cell = str(cell).replace('.', ',')
                    elif isinstance(cell, str):
                        cell = cell.replace('.', ',')
                    new_row.append(cell)
                writer.writerow(new_row)

        # Converter CSV de volta para Excel
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = aba_a_manter
        with open(temp_csv, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                ws_out.append(row)
        
        # Criar diretório se não existir
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        wb_out.save(output_path)

        # Remover arquivo temporário
        if os.path.exists(temp_csv):
            os.remove(temp_csv)
        
        debug_print(f"Multiplan processado: {output_path}")

    def get_row_values_from_cell(self, ws, cell_ref, novo_nome):
        """Extrai valores da linha a partir de uma célula"""
        col = openpyxl.utils.column_index_from_string(''.join(filter(str.isalpha, cell_ref)))
        row = int(''.join(filter(str.isdigit, cell_ref)))
        max_col = ws.max_column
        values = []
        for i in range(col, max_col + 1):
            value = ws.cell(row=row, column=i).value
            if i == col:
                values.append(novo_nome)
            else:
                values.append(value)
        return values

    def padroniza_trimestres(self, trimestre_raw):
        """Padroniza nomes dos trimestres"""
        trimestres = []
        for col in trimestre_raw:
            if isinstance(col, str):
                match = re.match(r"(\d)[TQº]?\s?(\d{2,4})", col, re.IGNORECASE)
                if match:
                    trimestre = f"{match.group(1)}Q{match.group(2)[-2:]}"
                    trimestres.append(trimestre)
                else:
                    trimestres.append(col)
            else:
                trimestres.append(col)
        return trimestres

    def align_trimestres(self, header, valores):
        """Alinha trimestres com valores"""
        return {t: v for t, v in zip(header, valores)}

    def sort_trimestres(self, trimestres):
        """Ordena trimestres cronologicamente"""
        def trimestre_key(t):
            m = re.match(r"(\d)Q(\d{2})", str(t))
            if m:
                trimestre = int(m.group(1))
                ano = int(m.group(2))
                ano += 2000 if ano < 80 else 1900
                return (ano, trimestre)
            return (9999, 9)
        return sorted(trimestres, key=trimestre_key)

    def round_val(self, val):
        """Arredonda valores"""
        try:
            f = float(str(val).replace(',', '.'))
            return round(f, 3)
        except (ValueError, TypeError):
            return val

    def to_percent_or_nan(self, val):
        try:
            num = float(str(val).replace(',', '.'))
            valor = round(num * 100, 2)
            valor_str = f"{valor:.2f}".replace('.', ',')
            return f"{valor_str}%"
        except (ValueError, TypeError):
            return "NaN"

    def process_files(self, progress_callback=None, status_callback=None):
        """Processa todos os arquivos"""
        try:
            # Executar inflation.py antes de tudo
            if status_callback:
                status_callback("Atualizando dados de inflação...")
            try:
                subprocess.run(['python', 'inflation.py'], check=True)
                debug_print("inflation.py executado com sucesso.")
            except Exception as e:
                debug_print(f"Erro ao executar inflation.py: {e}")
                if status_callback:
                    status_callback(f"Erro ao atualizar inflação: {e}")
            
            if status_callback:
                status_callback("Iniciando processamento dos arquivos...")
            
            # Definir os caminhos dos arquivos
            files_to_process = [
                {
                    'input': os.path.join('reports', 'Iguatemi Planilha 1T25.xlsx'),
                    'output': os.path.join('data_treated', 'iguatemi_data.xlsx'),
                    'function': self.treat_data_iguatemi,
                    'name': 'Iguatemi'
                },
                {
                    'input': os.path.join('reports', 'Allos Planilha 1T25.xlsx'),
                    'output': os.path.join('data_treated', 'allos_data.xlsx'),
                    'function': self.treat_data_allos,
                    'name': 'Allos'
                },
                {
                    'input': os.path.join('reports', 'Multiplan Planilha 1T25.xlsx'),
                    'output': os.path.join('data_treated', 'Multiplan_data.xlsx'),
                    'function': self.treat_data_multiplan,
                    'name': 'Multiplan'
                }
            ]
            
            # Verificar se todos os arquivos de entrada existem
            missing_files = []
            for file_info in files_to_process:
                if not os.path.exists(file_info['input']):
                    missing_files.append(file_info['input'])
            
            if missing_files:
                raise FileNotFoundError(f"Arquivos não encontrados: {', '.join(missing_files)}")
            
            # Processar cada arquivo
            for i, file_info in enumerate(files_to_process):
                if status_callback:
                    status_callback(f"Processando {file_info['name']}...")
                file_info['function'](file_info['input'], file_info['output'])
                if progress_callback:
                    progress_callback((i + 1) * 25)  # 25% por arquivo
            
            if status_callback:
                status_callback("Consolidando dados...")
            
            # Consolidar dados
            self.consolidate_data()
            
            if progress_callback:
                progress_callback(100)
            
            if status_callback:
                status_callback("Processamento concluído!")
            
            return True
            
        except Exception as e:
            debug_print(f"ERRO durante o processamento: {e}")
            if status_callback:
                status_callback(f"Erro: {str(e)}")
            logging.error("Erro crítico", exc_info=True)
            return False

    def consolidate_data(self):
        """Consolida os dados das três empresas"""
        empresas_info = [
            # (caminho, célula_inicial, nome_linha)
            (os.path.join("data_treated", "allos_data.xlsx"), "B17", "Trimestres Allos"),
            (os.path.join("data_treated", "allos_data.xlsx"), "B22", "SSS Allos"),
            (os.path.join("data_treated", "allos_data.xlsx"), "B14", "SSR Allos"),
            (os.path.join("data_treated", "allos_data.xlsx"), "B23", "OC Allos"),
            (os.path.join("data_treated", "allos_data.xlsx"), "B25", "TXOcup Allos"),
            (os.path.join("data_treated", "allos_data.xlsx"), "B24", "InadimplenciaLiq Allos"),
            
            (os.path.join("data_treated", "iguatemi_data.xlsx"), "B48", "Trimestres Iguatemi"),
            (os.path.join("data_treated", "iguatemi_data.xlsx"), "B54", "SSS Iguatemi"),
            (os.path.join("data_treated", "iguatemi_data.xlsx"), "B17", "SSR Iguatemi"),
            (os.path.join("data_treated", "iguatemi_data.xlsx"), "B19", "OC Iguatemi"),
            (os.path.join("data_treated", "iguatemi_data.xlsx"), "B20", "TXOcup Iguatemi"),
            (os.path.join("data_treated", "iguatemi_data.xlsx"), "B21", "InadimplenciaLiq Iguatemi"),

            (os.path.join("data_treated", "Multiplan_data.xlsx"), "B6", "Trimestres Multiplan"),
            (os.path.join("data_treated", "Multiplan_data.xlsx"), "B37", "SSS Multiplan"),
            (os.path.join("data_treated", "Multiplan_data.xlsx"), "B38", "SSR Multiplan"),
            (os.path.join("data_treated", "Multiplan_data.xlsx"), "B39", "OC Multiplan"),
            (os.path.join("data_treated", "Multiplan_data.xlsx"), "B43", "TXOcup Multiplan"),
            (os.path.join("data_treated", "Multiplan_data.xlsx"), "B45", "InadimplenciaLiq Multiplan")
            
            
        ]
        
        # Verificar se os arquivos existem
        arquivos_faltando = []
        for caminho, _, _ in empresas_info:
            if not os.path.exists(caminho):
                arquivos_faltando.append(caminho)
        
        if arquivos_faltando:
            raise FileNotFoundError(f"Arquivos não encontrados: {', '.join(set(arquivos_faltando))}")
        
        # Separar cabeçalhos (trimestres) e métricas por empresa
        headers = []
        metricas_linhas = []
        
        for path, cell_ref, nome_linha in empresas_info:
            debug_print(f"Processando: {nome_linha} de {path}")
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            valores = self.get_row_values_from_cell(ws, cell_ref, nome_linha)
            print(f"[DEBUG] Extraindo: {nome_linha} de {path} -> {valores}")  # Debug print
            if nome_linha.startswith("Trimestres"):
                headers.append((nome_linha, self.padroniza_trimestres(valores[1:])))
            else:
                metricas_linhas.append((nome_linha, valores[1:]))

            print("Trimestres Iguatemi:", valores[1:])
            print("Valores SSS Iguatemi:", valores[1:])

        # Montar dicionário: (empresa, metrica) -> {trimestre: valor}
        empresa_metrica_trimestres = {}
        empresa_trimestres = {}
        for (nome_header, trimestres) in headers:
            empresa = nome_header.replace("Trimestres ", "")
            empresa_trimestres[empresa] = trimestres

        for nome_linha, valores in metricas_linhas:
            # Junta tudo depois do primeiro espaço como nome da empresa
            split_idx = nome_linha.find(' ')
            if split_idx != -1:
                metrica = nome_linha[:split_idx]
                empresa = nome_linha[split_idx+1:].strip()
            else:
                metrica = nome_linha
                empresa = ""
            print(f"[DEBUG] Procurando trimestres para empresa: '{empresa}' em {list(empresa_trimestres.keys())}")
            trimestres = empresa_trimestres.get(empresa, [])
            empresa_metrica_trimestres[(empresa, metrica)] = self.align_trimestres(trimestres, valores)

        # Unir todos os trimestres possíveis, ordenados
        todos_trimestres = []
        for trimestres in empresa_trimestres.values():
            todos_trimestres.extend(trimestres)
        todos_trimestres = self.sort_trimestres(sorted(set(t for t in todos_trimestres if t not in (None, '', 'n/a', 'n.d.'))))

        # Criar nova planilha consolidada
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = "Consolidado"
        ws_out.append(["Empresa", "Métrica"] + todos_trimestres)
        for (empresa, metrica), valores_dict in empresa_metrica_trimestres.items():
            linha = [empresa, metrica]
            for t in todos_trimestres:
                val = self.to_percent_or_nan(valores_dict.get(t, ""))
                if val is None or str(val).strip().lower() in ["nan", "nan%", "none", "nan", "nan%", "nan%", ""]:
                    val = ''
                linha.append(val)
            ws_out.append(linha)

        # Salvar arquivo consolidado
        output_path = Path("data_treated") / "Consolidado.xlsx"
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb_out.save(str(output_path))
        debug_print(f"Arquivo consolidado salvo em: {output_path}")

        # Adicionar aba IPCA/IGPM
        ipca_igpm_csv = "ipca_igpm_20250701.csv"
        if os.path.exists(ipca_igpm_csv):
            self.add_ipca_igpm_sheet(str(output_path), ipca_igpm_csv)
        else:
            debug_print(f"Arquivo {ipca_igpm_csv} não encontrado. Aba IPCA_IGPM não adicionada.")

        # Inserir inflação trimestral na segunda aba
        self.insert_trimestral_inflation_to_second_sheet(str(output_path))
        self.insert_sss_ssr_descontado_in_first_sheet(str(output_path))

        # Gera e salva o gráfico automaticamente
        self.plot_all_metrics_from_excel(str(output_path))
        # Mover para área de trabalho
        desktop = Path.home() / "Desktop"
        dest = desktop / "Consolidado.xlsx"
        shutil.copy2(output_path, dest)
        debug_print(f"Arquivo consolidado copiado para: {dest}")
        return str(dest)

    def to_number_or_nan(self, val):
        try:
            # Tenta converter para float (aceita int também)
            return float(str(val).replace(',', '.'))
        except (ValueError, TypeError):
            return "NaN"

    def add_ipca_igpm_sheet(self, excel_path, csv_path=None):
        """Adiciona uma nova aba com os dados de IPCA/IGPM ao arquivo Excel consolidado."""
        import openpyxl
        import csv
        import glob
        import os
        wb = openpyxl.load_workbook(excel_path)
        # Remove a aba se já existir para evitar duplicidade
        if 'IPCA_IGPM' in wb.sheetnames:
            std = wb['IPCA_IGPM']
            wb.remove(std)
        ws = wb.create_sheet('IPCA_IGPM')
        # Buscar o CSV mensal mais recente
        csv_files = glob.glob('ipca_igpm_*.csv')
        csv_files = [f for f in csv_files if 'trimestre' not in f]
        if csv_files:
            csv_files.sort(reverse=True)
            csv_path = csv_files[0]
        else:
            ws.append(['Nenhum arquivo ipca_igpm_*.csv encontrado'])
            wb.save(excel_path)
            debug_print(f"Aba IPCA_IGPM adicionada ao arquivo: {excel_path}")
            return
        # Adiciona dados mensais, trocando ponto por vírgula nos valores
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                new_row = []
                for i, value in enumerate(row):
                    if i > 0 and value:
                        value = value.replace('.', ',')
                    new_row.append(value)
                ws.append(new_row)
        # Adiciona linha em branco
        ws.append([])
        # Adiciona dados trimestrais, se existir o arquivo
        trimestral_csv = 'ipca_igpm_trimestres.csv'
        if os.path.exists(trimestral_csv):
            with open(trimestral_csv, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    ws.append(row)
        else:
            ws.append(['Arquivo ipca_igpm_trimestres.csv não encontrado'])
        wb.save(excel_path)
        debug_print(f"Aba IPCA_IGPM adicionada ao arquivo: {excel_path}")

    def insert_trimestral_inflation_to_second_sheet(self, excel_path):
        import openpyxl
        import csv
        wb = openpyxl.load_workbook(excel_path)
        # Segunda aba
        if len(wb.sheetnames) < 2:
            debug_print('Menos de duas abas no consolidado, não foi possível inserir inflação trimestral.')
            return
        ws = wb.worksheets[1]
        # Ler dados trimestrais
        trimestral_csv = 'ipca_igpm_trimestres.csv'
        if not os.path.exists(trimestral_csv):
            debug_print('Arquivo ipca_igpm_trimestres.csv não encontrado para inserir na segunda aba.')
            return
        with open(trimestral_csv, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        # Inserir cabeçalho na coluna E (índice 5)
        for i, row in enumerate(rows):
            for j, value in enumerate(row):
                # Se for valor numérico, garantir vírgula como separador decimal
                if j > 0 and value:
                    value = value.replace('.', ',')
                ws.cell(row=i+1, column=5+j, value=value)
        wb.save(excel_path)
        debug_print('Inflação trimestral inserida na coluna E da segunda aba.')

    def insert_sss_ssr_descontado_in_first_sheet(self, excel_path):
        import openpyxl
        import csv
        import os
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.worksheets[0]
        # Ler cabeçalho e trimestres
        header = [cell.value for cell in ws[1]]
        trimestres = header[2:]
        # Ler inflação trimestral
        trimestral_csv = 'ipca_igpm_trimestres.csv'
        if not os.path.exists(trimestral_csv):
            debug_print('Arquivo ipca_igpm_trimestres.csv não encontrado para inserir SSS/SSR descontado.')
            return
        with open(trimestral_csv, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            inflacao = {row['trimestre'].replace('T','Q'): {'IPCA': row['IPCA'].replace(',', '.') if row['IPCA'] else None,
                                                            'IGPM': row['IGPM'].replace(',', '.') if row['IGPM'] else None}
                        for row in reader}
        # Coletar linhas SSS e SSR
        sss_rows = []
        ssr_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            empresa, metrica, *valores = row
            if metrica == 'SSS':
                sss_rows.append((empresa, valores))
            elif metrica == 'SSR':
                ssr_rows.append((empresa, valores))
        # Calcular linhas descontadas
        new_rows = []
        for empresa, valores in sss_rows:
            new_row = [empresa, 'SSS_Descontado']
            for i, t in enumerate(trimestres):
                t_pad = str(t).replace('T', 'Q')
                val = valores[i]
                if val is None or str(val).strip().lower() in ['nan', 'nan%','none','nan','nan%','nan%']:
                    new_row.append('')
                    continue
                try:
                    val_f = float(str(val).replace(',', '.').replace('%',''))
                except:
                    new_row.append('')
                    continue
                fator = inflacao.get(t_pad, {}).get('IPCA')
                if fator is None or fator == '' or fator == 'nan':
                    new_row.append('')
                    continue
                try:
                    fator_f = float(fator)
                except:
                    new_row.append('')
                    continue
                try:
                    descontado = (1+val_f/100)/(1+fator_f/100)-1
                    descontado = descontado*100
                    descontado_str = f"{descontado:.2f}".replace('.', ',') + '%'
                    if descontado_str.lower() in ['nan%', 'nan', 'none', 'nan%']:
                        descontado_str = ''
                    new_row.append(descontado_str)
                except:
                    new_row.append('')
            new_rows.append(new_row)
        for empresa, valores in ssr_rows:
            new_row = [empresa, 'SSR_Descontado']
            for i, t in enumerate(trimestres):
                t_pad = str(t).replace('T', 'Q')
                val = valores[i]
                if val is None or str(val).strip().lower() in ['nan', 'nan%','none','nan','nan%','nan%']:
                    new_row.append('')
                    continue
                try:
                    val_f = float(str(val).replace(',', '.').replace('%',''))
                except:
                    new_row.append('')
                    continue
                fator = inflacao.get(t_pad, {}).get('IGPM')
                if fator is None or fator == '' or fator == 'nan':
                    new_row.append('')
                    continue
                try:
                    fator_f = float(fator)
                except:
                    new_row.append('')
                    continue
                try:
                    descontado = (1+val_f/100)/(1+fator_f/100)-1
                    descontado = descontado*100
                    descontado_str = f"{descontado:.2f}".replace('.', ',') + '%'
                    if descontado_str.lower() in ['nan%', 'nan', 'none', 'nan%']:
                        descontado_str = ''
                    new_row.append(descontado_str)
                except:
                    new_row.append('')
            new_rows.append(new_row)
        # Inserir a partir da linha 18 (A18)
        for idx, row in enumerate(new_rows):
            for col_idx, value in enumerate(row, 1):
                if value is None or str(value).strip().lower() in ['nan', 'nan%','none','nan','nan%','nan%']:
                    value = ''
                ws.cell(row=18+idx, column=col_idx, value=value)
        wb.save(excel_path)
        debug_print('SSS e SSR descontados inseridos na primeira aba a partir de A18.')

    def plot_all_metrics_from_excel(self, excel_path, output_path='graficos_consolidado.png'):
        import matplotlib.pyplot as plt
        import pandas as pd
        import numpy as np
        import os
        import sys

        # Lê a primeira aba do consolidado
        df = pd.read_excel(excel_path, sheet_name=0)
        metricas = ['SSS', 'SSR', 'OC', 'TXOcup', 'InadimplenciaLiq']
        metricas_existentes = df['Métrica'].unique()
        metricas_para_plotar = [m for m in metricas if m in metricas_existentes]
        if not metricas_para_plotar:
            return

        trimestres_todos = df.columns[2:].tolist()
        trimestres_filtrados = []
        for trimestre in trimestres_todos:
            if isinstance(trimestre, str) and 'Q' in trimestre:
                ano_str = trimestre.split('Q')[-1]
                try:
                    ano = int(ano_str)
                    if ano >= 20:
                        trimestres_filtrados.append(trimestre)
                except:
                    continue

        if not trimestres_filtrados:
            return

        n_metricas = len(metricas_para_plotar)
        n_cols = min(3, n_metricas)
        n_rows = (n_metricas + n_cols - 1) // n_cols
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 5*n_rows))
        fig.suptitle('Métricas Consolidadas - 2020 em diante', fontsize=16, fontweight='bold')
        if n_metricas == 1:
            axes = [axes]
        elif n_rows == 1:
            axes = [axes] if n_cols == 1 else axes
        else:
            axes = axes.flatten()
        cores = ['#1f77b4', '#ff7f0e', '#2ca02c']

        # Encontrar índices dos 1Q de cada ano para os ticks
        xticks = []
        xticklabels = []
        anos_vistos = set()
        for j, trimestre in enumerate(trimestres_filtrados):
            if isinstance(trimestre, str) and trimestre.startswith('1Q'):
                ano_str = trimestre.split('Q')[-1]
                try:
                    ano = int(ano_str)
                    if ano not in anos_vistos:
                        xticks.append(j)
                        xticklabels.append(trimestre)
                        anos_vistos.add(ano)
                except:
                    continue

        for idx, metrica in enumerate(metricas_para_plotar):
            if idx >= len(axes):
                break
            ax = axes[idx]
            dados_metrica = df[df['Métrica'] == metrica]
            if not dados_metrica.empty:
                for i, (_, row) in enumerate(dados_metrica.iterrows()):
                    empresa = str(row['Empresa'])
                    valores = []
                    for trimestre in trimestres_filtrados:
                        valor = row[trimestre]
                        if pd.isna(valor) or valor == '' or str(valor).lower() in ['nan', 'nan%']:
                            valores.append(np.nan)
                        else:
                            try:
                                valor_str = str(valor).replace('%', '').replace(',', '.')
                                valores.append(float(valor_str))
                            except:
                                valores.append(np.nan)
                    print(f"Plotando {metrica} - {empresa}: {valores}")  # DEBUG
                    ax.plot(trimestres_filtrados, valores, linewidth=1.0, label=empresa, color=cores[i % len(cores)], marker=None)
            ax.set_title(f'{metrica}', fontweight='bold', fontsize=12)
            ax.set_xlabel('Trimestre')
            ax.set_ylabel('%')
            ax.legend()
            ax.grid(True, alpha=0.3)
            # Mostrar apenas 1Q de cada ano no eixo X
            ax.set_xticks([trimestres_filtrados[i] for i in xticks])
            ax.set_xticklabels(xticklabels, rotation=45)
            ax.autoscale_view()
        for idx in range(len(metricas_para_plotar), len(axes)):
            fig.delaxes(axes[idx])
        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()
        print(f"Gráfico salvo em: {output_path}")
        # Abrir a imagem automaticamente
        try:
            if sys.platform.startswith('win'):
                os.startfile(output_path)
            elif sys.platform.startswith('darwin'):
                import subprocess
                subprocess.run(['open', output_path])
            else:
                import subprocess
                subprocess.run(['xdg-open', output_path])
            print(f"Imagem aberta: {output_path}")
        except Exception as e:
            print(f"Erro ao abrir imagem: {e}")

def run_processing(progress, status_label, root):
    """Executa o processamento em thread separada"""
    try:
        processor = ReportProcessor()
        
        def progress_callback(value):
            progress["value"] = value
            root.update_idletasks()
        
        def status_callback(message):
            status_label.config(text=message)
            root.update_idletasks()
        
        success = processor.process_files(progress_callback, status_callback)
        
        if success:
            # Abrir arquivo Excel
            desktop_path = Path.home() / "Desktop"
            consolidado_path = desktop_path / "Consolidado.xlsx"
            
            if consolidado_path.exists():
                try:
                    if sys.platform.startswith('win'):
                        os.startfile(str(consolidado_path))
                    elif sys.platform.startswith('darwin'):
                        import subprocess
                        subprocess.run(['open', str(consolidado_path)])
                    else:
                        import subprocess
                        subprocess.run(['xdg-open', str(consolidado_path)])
                    
                    messagebox.showinfo("Sucesso", 
                        f"Processamento concluído!\nArquivo salvo em:\n{consolidado_path}\nO arquivo foi aberto automaticamente.")
                except Exception as e:
                    debug_print(f"Erro ao abrir arquivo: {e}")
                    messagebox.showinfo("Sucesso", 
                        f"Processamento concluído!\nArquivo salvo em:\n{consolidado_path}\nAbra o arquivo manualmente.")
            else:
                messagebox.showwarning("Aviso", f"Arquivo não encontrado: {consolidado_path}")
        else:
            messagebox.showerror("Erro", "Falha no processamento. Verifique os logs.")
            
    except Exception as e:
        debug_print(f"ERRO CRÍTICO na interface: {e}")
        debug_print(f"Traceback: {traceback.format_exc()}")
        logging.error("Erro crítico na interface", exc_info=True)
        # ... resto do código
    finally:
        debug_print("=== PROCESSAMENTO FINALIZADO ===")

def select_and_copy_files(root, status_callback):
    """Permite ao usuário selecionar os 3 arquivos Excel"""
    reports_dir = Path("reports")
    
    try:
        reports_dir.mkdir(exist_ok=True)
        debug_print(f"Pasta criada: {reports_dir}")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao criar pasta reports: {e}")
        return False

    file_mappings = [
        ("Selecione o arquivo da ALLOS", "Allos Planilha 1T25.xlsx"),
        ("Selecione o arquivo do IGUATEMI", "Iguatemi Planilha 1T25.xlsx"),
        ("Selecione o arquivo da MULTIPLAN", "Multiplan Planilha 1T25.xlsx"),
    ]
    
    for i, (title, reports_name) in enumerate(file_mappings, 1):
        status_callback(f"Selecionando arquivo {i}/3: {title}")
        root.update_idletasks()
        
        file_path = filedialog.askopenfilename(
            title=title,
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("Todos os arquivos", "*.*")
            ],
            parent=root
        )
        
        if not file_path:
            messagebox.showwarning("Seleção Cancelada", f"Nenhum arquivo foi selecionado para: {title}")
            return False
        
        try:
            reports_dest = reports_dir / reports_name
            shutil.copy2(file_path, reports_dest)
            debug_print(f"Copiado para reports: {file_path} -> {reports_dest}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao copiar arquivo:\n{file_path}\n\nErro: {e}")
            return False
    
    debug_print("Todos os arquivos foram copiados com sucesso")
    return True

def main():
    """Função principal da interface gráfica"""
    try:
        debug_print("Iniciando aplicação...")
        
        # Configurar janela principal
        root = tk.Tk()
        root.title("Consolidador de Dados - Automatização de Reports")
        root.geometry("540x420")
        root.resizable(False, False)
        
        # Centralizar janela
        try:
            root.eval('tk::PlaceWindow . center')
        except:
            # Fallback para centralizar manualmente
            root.update_idletasks()
            x = (root.winfo_screenwidth() // 2) - (540 // 2)
            y = (root.winfo_screenheight() // 2) - (420 // 2)
            root.geometry(f"540x420+{x}+{y}")

        # Frame principal
        main_frame = tk.Frame(root, bg="#f8f9fa", padx=30, pady=30)
        main_frame.pack(fill="both", expand=True)

        # Título
        title_label = tk.Label(main_frame, text="Consolidador de Dados", font=("Arial", 20, "bold"), fg="#2c3e50", bg="#f8f9fa")
        title_label.pack(pady=(0, 8))

        # Subtítulo
        subtitle_label = tk.Label(main_frame, text="Allos   •   Iguatemi   •   Multiplan", font=("Arial", 12, "italic"), fg="#2980b9", bg="#f8f9fa")
        subtitle_label.pack(pady=(0, 18))

        # Instruções
        instruction_label = tk.Label(main_frame, text="Selecione os arquivos Excel das empresas e clique em Iniciar.", font=("Arial", 11), fg="#34495e", bg="#f8f9fa")
        instruction_label.pack(pady=(0, 10))

        # Status
        status_label = tk.Label(main_frame, text="Aguardando ação do usuário...", font=("Arial", 10, "italic"), fg="#16a085", bg="#f8f9fa")
        status_label.pack(pady=(0, 18))

        # Barra de progresso
        progress = ttk.Progressbar(main_frame, orient="horizontal", length=400, mode="determinate")
        progress.pack(pady=(0, 25))
        progress["value"] = 0

        def start_process():
            """Inicia o processo de seleção de arquivos e processamento"""
            debug_print("=== INICIANDO PROCESSO ===")
            start_btn.config(state="disabled")
            progress["value"] = 0
            status_label.config(text="Selecione os arquivos conforme solicitado...")
            root.update_idletasks()
            if not select_and_copy_files(root, lambda msg: status_label.config(text=msg)):
                status_label.config(text="Processo cancelado")
                start_btn.config(state="normal")
                return
            progress["value"] = 5
            status_label.config(text="Arquivos selecionados. Iniciando processamento...")
            root.update_idletasks()
            thread = threading.Thread(target=run_processing, args=(progress, status_label, root), daemon=True)
            thread.start()

        # Botão principal (garantir que sempre aparece)
        start_btn = tk.Button(
            main_frame,
            text="INICIAR PROCESSAMENTO",
            command=start_process,
            font=("Arial", 13, "bold"),
            bg="#27ae60",
            fg="white",
            padx=40,
            pady=18,
            relief="flat",
            cursor="hand2",
            activebackground="#219150",
            activeforeground="white"
        )
        start_btn.pack(pady=(0, 30))

        # Rodapé
        footer = tk.Frame(main_frame, bg="#f8f9fa")
        footer.pack(side="bottom", fill="x", pady=(30, 0))
        version_label = tk.Label(footer, text="v2.0 - Automatização de Reports", font=("Arial", 8), fg="#7f8c8d", bg="#f8f9fa")
        version_label.pack(side="left")
        credits_label = tk.Label(footer, text="Lev Asset Management ©", font=("Arial", 8), fg="#7f8c8d", bg="#f8f9fa")
        credits_label.pack(side="right")

        # Iniciar interface
        debug_print("Interface iniciada com sucesso")
        root.mainloop()
        
    except Exception as e:
        debug_print(f"ERRO CRÍTICO na interface: {e}")
        debug_print(f"Traceback: {traceback.format_exc()}")
        logging.error("Erro crítico na interface", exc_info=True)
        
        # Tentar mostrar erro em uma janela simples
        try:
            error_root = tk.Tk()
            error_root.withdraw()  # Esconder janela principal
            messagebox.showerror("Erro Crítico", 
                f"Erro ao iniciar a interface:\n{str(e)}\n\nVerifique se o Python e tkinter estão instalados corretamente.")
            error_root.destroy()
        except:
            print("ERRO: Não foi possível mostrar a interface gráfica.")
            print("Verifique se o tkinter está instalado e funcionando.")
            input("Pressione Enter para sair...")

def add_sss_allos_chart(input_excel, output_excel):
    # Lê a primeira aba do consolidado
    df = pd.read_excel(input_excel, sheet_name=0)
    
    # Filtra apenas a linha de SSS da Allos
    sss_allos = df[(df['Empresa'] == 'Allos') & (df['Métrica'] == 'SSS')]
    if sss_allos.empty:
        print("Linha de SSS da Allos não encontrada.")
        return
    
    # Cria um novo arquivo Excel com XlsxWriter
    with pd.ExcelWriter(output_excel, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Consolidado', index=False)
        workbook  = writer.book
        worksheet = writer.sheets['Consolidado']
        
        # Descobre o range dos dados
        n_trimestres = len(df.columns) - 2
        row = sss_allos.index[0]
        
        # Adiciona gráfico de linhas para SSS Allos
        chart = workbook.add_chart({'type': 'line'})
        chart.add_series({
            'name':       ['Consolidado', row+1, 0],
            'categories': ['Consolidado', 0, 2, 0, n_trimestres+1],
            'values':     ['Consolidado', row+1, 2, row+1, n_trimestres+1],
        })
        chart.set_title({'name': 'SSS Allos'})
        chart.set_x_axis({'name': 'Trimestre'})
        chart.set_y_axis({'name': '%'})
        worksheet.insert_chart('A2', chart)
        
    print(f"Arquivo com gráfico SSS Allos salvo em: {output_excel}")

# Teste simples
if __name__ == "__main__":
    main() 
    